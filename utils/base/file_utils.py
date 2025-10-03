import asyncio
import time
import aiofiles

from pathlib import Path
from aiocsv import AsyncWriter
from loguru import logger

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from models import ModuleType, OperationResult, Account


class FileOperations:
    def __init__(self, base_path: str = "./results"):
        self.base_path = Path(base_path)
        self.lock = asyncio.Lock()
        self.module_paths: dict[ModuleType, dict[str, Path]] = {
            # "统计": {
            #     "基础": self.base_path / "统计" / "accounts_stats.xlsx",
            # },
            "request_tokens": {
                "success": self.base_path / "request_tokens" / "request_tokens_success.txt",
                "failed": self.base_path / "request_tokens" / "request_tokens_failed.txt",
            },
            "top_up_game_balance": {
                "success": self.base_path / "top_up_game_balance" / "top_up_game_balance_success.txt",
                "failed": self.base_path / "top_up_game_balance" / "top_up_game_balance_failed.txt",
            },
            "play_games": {
                "success": self.base_path / "play_games" / "play_games_success.txt",
                "failed": self.base_path / "play_games" / "play_games_failed.txt",
            },
            "mint_nft": {
                "success": self.base_path / "mint_nft" / "mint_nft_success.txt",
                "failed": self.base_path / "mint_nft" / "mint_nft_failed.txt",
            },
            "all_in_one": {
                "success": self.base_path / "all_in_one" / "all_in_one_success.txt",
                "failed": self.base_path / "all_in_one" / "all_in_one_failed.txt",
            }
        }

    async def setup_files(self):
        self.base_path.mkdir(exist_ok=True)
        for module_name, module_paths in self.module_paths.items():
            for path_key, path in module_paths.items():
                path.parent.mkdir(parents=True, exist_ok=True)
                if module_name == "stats":
                    continue
                else:
                    path.touch(exist_ok=True)

    async def setup_stats(self):
        self.base_path.mkdir(exist_ok=True)

        for module_name, module_paths in self.module_paths.items():
            if module_name == "stats":
                timestamp = int(time.time())
                for path_key, path in module_paths.items():
                    path.parent.mkdir(parents=True, exist_ok=True)
                    if path_key == "base":
                        new_path = path.parent / f"accounts_stats_{timestamp}.xlsx"
                        self.module_paths[module_name][path_key] = new_path
                        await asyncio.to_thread(self._create_excel_with_header, new_path)

    @staticmethod
    def _create_excel_with_header(xlsx_path: Path, sheet_name: str = "统计"):
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = sheet_name
        ws.append([
            "钱包地址",
            "私钥",
            "每日积分",
            "总积分",
            "邀请码",
            "是否连接 Twitter",
            "Twitter 用户名",
        ])

        widths = [46, 66, 14, 14, 14, 20, 24]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(xlsx_path)

    async def _append_excel_row(self, xlsx_path: Path, row: list, sheet_name: str = "Stats"):
        await asyncio.to_thread(self._append_excel_row_sync, xlsx_path, row, sheet_name)

    @staticmethod
    def _append_excel_row_sync(xlsx_path: Path, row: list, sheet_name: str = "Stats"):
        wb = load_workbook(xlsx_path)
        ws: Worksheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        ws.append(row)
        wb.save(xlsx_path)
        wb.close()

    @staticmethod
    def _desensitize_private_key(private_key: str) -> str:
        if not private_key or len(private_key) < 10:
            return "[REDACTED_KEY]"
        return f"{private_key[:5]}...{private_key[-5:]}"

    async def export_result(self, result: OperationResult, module: ModuleType):
        if module not in self.module_paths:
            raise ValueError(f"Unknown module: {module}")

        file_path = self.module_paths[module]["success" if result["status"] else "failed"]
        async with self.lock:
            try:
                async with aiofiles.open(file_path, "a") as file:
                    desensitized_pk = self._desensitize_private_key(result['pk_or_mnemonic'])
                    await file.write(f"{desensitized_pk}\n")
            except IOError as e:
                logger.error(f"账户: {result['pk_or_mnemonic']} | 写入文件错误 (IOError): {e}")
            except Exception as e:
                logger.error(f"账户: {result['pk_or_mnemonic']} | 写入文件错误: {e}")

    async def export_stats(self, result: OperationResult):
        file_path = self.module_paths["stats"]["base"]
        async with self.lock:
            try:
                if result["status"] is True:
                    row = [
                        result["data"]["wallet_address"],
                        self._desensitize_private_key(result["data"]["private_key"]),
                        result["data"]["daily_points"],
                        result["data"]["total_points"],
                        result["data"]["invite_code"],
                        result["data"]["user_info"]["linkTwitter"],
                        result["data"]["user_info"]["xUsername"],
                    ]
                else:
                    row = [
                        result["data"]["wallet_address"],
                        self._desensitize_private_key(result["data"]["private_key"]),
                        "N/A",
                        "N/A",
                        "N/A",
                        "N/A",
                        "N/A",
                    ]

                await self._append_excel_row(file_path, row)

            except IOError as e:
                logger.error(f"账户: {result['data']['wallet_address']} | 写入 Excel 错误 (IOError): {e}")
            except Exception as e:
                logger.error(f"账户: {result['data']['wallet_address']} | 写入 Excel 错误: {e}")
